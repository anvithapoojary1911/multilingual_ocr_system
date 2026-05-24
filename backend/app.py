from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pytesseract
from PIL import Image, ImageFilter, ImageEnhance
import cv2
import numpy as np
import io
import os
import base64
import json
import time
import zipfile
import tempfile
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

LANGUAGE_MAP = {
    'english': 'eng',
    'hindi': 'hin',
    'kannada': 'kan',
    'tamil': 'tam',
    'telugu': 'tel',
    'marathi': 'mar',
    'bengali': 'ben',
    'gujarati': 'guj',
    'punjabi': 'pan',
    'malayalam': 'mal',
    'english+hindi': 'eng+hin',
    'english+kannada': 'eng+kan',
    'english+hindi+kannada': 'eng+hin+kan',
    'all': 'eng+hin+kan+tam+tel',
}

def preprocess_image(image_bytes, preprocessing_level='standard'):
    """Preprocess image for better OCR accuracy."""
    nparr = np.frombuffer(image_bytes, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

    if img is None:
        raise ValueError("Could not decode image")

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    if preprocessing_level == 'minimal':
        processed = gray

    elif preprocessing_level == 'standard':
        # Denoise
        denoised = cv2.fastNlMeansDenoising(gray, h=10)
        # Adaptive threshold
        processed = cv2.adaptiveThreshold(
            denoised, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY, 11, 2
        )

    elif preprocessing_level == 'aggressive':
        # Denoise
        denoised = cv2.fastNlMeansDenoising(gray, h=15)
        # Sharpen
        kernel = np.array([[-1,-1,-1], [-1,9,-1], [-1,-1,-1]])
        sharpened = cv2.filter2D(denoised, -1, kernel)
        # Adaptive threshold
        thresholded = cv2.adaptiveThreshold(
            sharpened, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY, 11, 2
        )
        # Morphological operations
        kernel_morph = np.ones((1, 1), np.uint8)
        processed = cv2.morphologyEx(thresholded, cv2.MORPH_CLOSE, kernel_morph)

    elif preprocessing_level == 'deskew':
        denoised = cv2.fastNlMeansDenoising(gray, h=10)
        # Deskew
        coords = np.column_stack(np.where(denoised > 0))
        if len(coords) > 0:
            angle = cv2.minAreaRect(coords)[-1]
            if angle < -45:
                angle = -(90 + angle)
            else:
                angle = -angle
            (h, w) = denoised.shape[:2]
            center = (w // 2, h // 2)
            M = cv2.getRotationMatrix2D(center, angle, 1.0)
            denoised = cv2.warpAffine(denoised, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
        processed = cv2.adaptiveThreshold(denoised, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
    else:
        processed = gray

    _, buffer = cv2.imencode('.png', processed)
    return buffer.tobytes()


def run_ocr(image_bytes, lang_code, psm=3, oem=3):
    """Run Tesseract OCR on image bytes."""
    pil_image = Image.open(io.BytesIO(image_bytes))
    config = f'--psm {psm} --oem {oem}'
    text = pytesseract.image_to_string(pil_image, lang=lang_code, config=config)
    return text.strip()


def get_ocr_confidence(image_bytes, lang_code):
    """Get word-level confidence data."""
    try:
        pil_image = Image.open(io.BytesIO(image_bytes))
        data = pytesseract.image_to_data(pil_image, lang=lang_code, output_type=pytesseract.Output.DICT)
        confidences = [c for c in data['conf'] if c != -1]
        avg_confidence = sum(confidences) / len(confidences) if confidences else 0
        return round(avg_confidence, 2)
    except:
        return 0


@app.route('/api/health', methods=['GET'])
def health():
    """Health check endpoint."""
    try:
        version = pytesseract.get_tesseract_version()
        return jsonify({'status': 'ok', 'tesseract_version': str(version)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/languages', methods=['GET'])
def get_languages():
    """Get available languages."""
    try:
        available = pytesseract.get_languages()
        return jsonify({'languages': list(LANGUAGE_MAP.keys()), 'installed': available})
    except Exception as e:
        return jsonify({'languages': list(LANGUAGE_MAP.keys()), 'installed': [], 'error': str(e)})


@app.route('/api/ocr', methods=['POST'])
def process_single_image():
    """Process a single image for OCR."""
    start_time = time.time()

    if 'image' not in request.files and 'image_base64' not in request.json if request.is_json else True:
        if 'image' not in request.files:
            return jsonify({'error': 'No image provided'}), 400

    language = request.form.get('language', 'english')
    preprocessing = request.form.get('preprocessing', 'standard')
    psm = int(request.form.get('psm', 3))

    lang_code = LANGUAGE_MAP.get(language, 'eng')

    if 'image' in request.files:
        file = request.files['image']
        image_bytes = file.read()
        filename = file.filename
    else:
        return jsonify({'error': 'No image file provided'}), 400

    try:
        # Preprocess
        processed_bytes = preprocess_image(image_bytes, preprocessing)

        # Run OCR
        text = run_ocr(processed_bytes, lang_code, psm)

        # Get confidence
        confidence = get_ocr_confidence(processed_bytes, lang_code)

        # Word count
        words = [w for w in text.split() if w.strip()]
        lines = [l for l in text.split('\n') if l.strip()]

        elapsed = round(time.time() - start_time, 2)

        return jsonify({
            'success': True,
            'text': text,
            'confidence': confidence,
            'word_count': len(words),
            'line_count': len(lines),
            'char_count': len(text),
            'language': language,
            'preprocessing': preprocessing,
            'processing_time': elapsed,
            'filename': filename
        })

    except Exception as e:
        return jsonify({'error': str(e), 'success': False}), 500


@app.route('/api/ocr/bulk', methods=['POST'])
def process_bulk_images():
    """Process multiple images."""
    start_time = time.time()

    if 'images' not in request.files:
        return jsonify({'error': 'No images provided'}), 400

    files = request.files.getlist('images')
    language = request.form.get('language', 'english')
    preprocessing = request.form.get('preprocessing', 'standard')
    export_format = request.form.get('export_format', 'json')

    lang_code = LANGUAGE_MAP.get(language, 'eng')
    results = []

    for file in files:
        file_start = time.time()
        try:
            image_bytes = file.read()
            processed_bytes = preprocess_image(image_bytes, preprocessing)
            text = run_ocr(processed_bytes, lang_code)
            confidence = get_ocr_confidence(processed_bytes, lang_code)
            words = [w for w in text.split() if w.strip()]

            results.append({
                'filename': file.filename,
                'text': text,
                'confidence': confidence,
                'word_count': len(words),
                'char_count': len(text),
                'processing_time': round(time.time() - file_start, 2),
                'status': 'success'
            })
        except Exception as e:
            results.append({
                'filename': file.filename,
                'text': '',
                'error': str(e),
                'status': 'error',
                'processing_time': round(time.time() - file_start, 2)
            })

    total_time = round(time.time() - start_time, 2)
    success_count = sum(1 for r in results if r['status'] == 'success')
    avg_confidence = sum(r.get('confidence', 0) for r in results if r['status'] == 'success')
    avg_confidence = round(avg_confidence / success_count, 2) if success_count > 0 else 0

    # Export
    export_path = None
    if export_format == 'excel':
        export_path = export_to_excel(results)
    elif export_format == 'txt':
        export_path = export_to_txt(results)

    response_data = {
        'success': True,
        'total_images': len(files),
        'successful': success_count,
        'failed': len(files) - success_count,
        'average_confidence': avg_confidence,
        'total_processing_time': total_time,
        'results': results
    }

    if export_path:
        response_data['export_file'] = os.path.basename(export_path)

    return jsonify(response_data)


def export_to_excel(results):
    """Export results to Excel."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'ocr_results_{timestamp}.xlsx'
    filepath = os.path.join(OUTPUT_FOLDER, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OCR Results"

    headers = ['#', 'Filename', 'Extracted Text', 'Confidence (%)', 'Words', 'Characters', 'Time (s)', 'Status']
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for i, result in enumerate(results, 2):
        ws.cell(row=i, column=1, value=i-1)
        ws.cell(row=i, column=2, value=result.get('filename', ''))
        ws.cell(row=i, column=3, value=result.get('text', ''))
        ws.cell(row=i, column=4, value=result.get('confidence', 0))
        ws.cell(row=i, column=5, value=result.get('word_count', 0))
        ws.cell(row=i, column=6, value=result.get('char_count', 0))
        ws.cell(row=i, column=7, value=result.get('processing_time', 0))
        ws.cell(row=i, column=8, value=result.get('status', ''))

    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['B'].width = 30
    wb.save(filepath)
    return filepath


def export_to_txt(results):
    """Export results to text file."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'ocr_results_{timestamp}.txt'
    filepath = os.path.join(OUTPUT_FOLDER, filename)

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(f"Multilingual OCR Results - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 60 + "\n\n")
        for i, result in enumerate(results, 1):
            f.write(f"[{i}] File: {result.get('filename', 'Unknown')}\n")
            f.write(f"Status: {result.get('status', 'unknown')}\n")
            f.write(f"Confidence: {result.get('confidence', 0)}%\n")
            f.write(f"Words: {result.get('word_count', 0)} | Chars: {result.get('char_count', 0)}\n")
            f.write(f"Text:\n{result.get('text', '')}\n")
            f.write("-" * 40 + "\n\n")
    return filepath


@app.route('/api/download/<filename>', methods=['GET'])
def download_file(filename):
    """Download an exported file."""
    filepath = os.path.join(OUTPUT_FOLDER, filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return jsonify({'error': 'File not found'}), 404


@app.route('/api/preprocess/preview', methods=['POST'])
def preview_preprocessing():
    """Return preprocessed image as base64 for preview."""
    if 'image' not in request.files:
        return jsonify({'error': 'No image provided'}), 400

    file = request.files['image']
    preprocessing = request.form.get('preprocessing', 'standard')

    try:
        image_bytes = file.read()
        processed_bytes = preprocess_image(image_bytes, preprocessing)

        # Convert to base64
        b64 = base64.b64encode(processed_bytes).decode('utf-8')
        return jsonify({'success': True, 'image': f'data:image/png;base64,{b64}'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
